import shodan
from openpyxl import load_workbook
from openpyxl import Workbook

class SearchData:
    def __init__(self):
        self.ip    = "-"
        self.ports = "-"
        self.os    = "-"
        self.last_update = "-"
        self.vulns = "-"
        self.country_name = "-"
        self.city = "-"
        self.organization = "-"
        self.data = None

def search_ip(SHODAN_API_KEY, SEARCH_IP):
    try:
        api = shodan.Shodan(SHODAN_API_KEY)
        host = api.host(SEARCH_IP)
        data = SearchData()
        result_list = []
        #
        for key in host:
            if key == 'ip_str':
                data.ip = host[key]
            elif key == 'ports':
                data.ports = host[key]
            elif key == 'os':
                data.os = host[key]
            elif key == 'last_update':
                data.last_update = host[key]
            elif key == 'vulns':
                data.vulns = host[key]
            elif key == 'country_name':
                data.country_name = host[key]
            elif key == 'city':
                data.city = host[key]
            elif key == 'org':
                data.organization = host[key]
            elif key == 'data':
                data.data = host[key]
        
        port = ''
        for p in data.ports :
            port += p + ','

        result_list.append([data.ip, data.city, data.country_name, data.organization, data.last_update, data.os, port, data.vulns])

        print('----------------------요약-----------------------')
        print('---------------------------------------------------')
        print("\tIp              :", data.ip)
        print("\tCity            :", data.city)
        print("\tCountry         :", data.country_name)
        print("\tOrganization    :", data.organization)
        print("\tLast Update     :", data.last_update)
        print("\tOs              :", data.os)
        print("\tPorts           :", data.ports)
        print("\tVulnerabilities :", data.vulns)
        print('---------------------------------------------------')
        print()
        print()

        
        print('----------------------디테일-----------------------')
        print('---------------------------------------------------')
        for item in data.data:
            print()
            print("\tPort                     :", item['port'])
            try:
                print("\tProtocol             :", item['transport'])
            except:
                print("\tProtocol             : - ")
            try:
                print("\tProgram              :", item['product'])
            except:
                print("\tProgram              :  - ")
            try:
                print("\tVulnerability(ies)   :", item['vulns'])
            except:
                print("\tVulnerability(ies)   :  - ")
        
        return result_list
    except:
        print("에러 발생")

if __name__ == '__main__':
    SHODAN_API_KEY = ''
    load_wb = load_workbook("./IP_LIST.xlsx", data_only=True)
    load_ws = load_wb['Sheet']

    write_wb = Workbook()
    write_ws = write_wb.active

    line = 1
    SEARCH_IP = '0.0.0.0'

    while SEARCH_IP != None :
        SEARCH_IP = load_ws.cell(line, 1).value                 #엑셀 입력
        line += 1
        seach_result = search_ip(SHODAN_API_KEY, SEARCH_IP)     #엑셀 읽어오기
        seach_result.insert(0, SEARCH_IP)                       #첫번째 셀에 IP 넣기
        
        write_ws.append(seach_result)
        write_wb.save('./IP_LIST.xlsx')                         #엑셀에 행으로 저장
        write_wb.close()
    print("----작업 완료----")